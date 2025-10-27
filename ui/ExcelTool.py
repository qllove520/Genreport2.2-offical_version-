import sys
import os
import shutil
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton,
    QVBoxLayout, QHBoxLayout, QFileDialog, QMessageBox, QTextEdit
)
from PyQt5.QtCore import Qt
from core.excel_utils import find_row_by_fuzzy_column_value,write_to_target_sheet
from openpyxl import load_workbook
from config.settings import Reference_msg_path,Testreport_path
from core.settings_manager import SettingsManager

class ExcelTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("项目台账写入工具")
        self.setMinimumSize(700, 500)

        self.settings_manager = SettingsManager()
        # 优先读取配置文件
        loaded_settings = self.settings_manager.load_settings(
            "acceptance_filling",
            default_settings={
                "data_file": Reference_msg_path,
                "template_file": Testreport_path
            }
        )
        self.data_file = loaded_settings.get("data_file", Reference_msg_path)
        self.template_file = loaded_settings.get("template_file", Testreport_path)

        # 新增输入字段变量
        self.input_fields = {}

        self.init_ui()

    def init_ui(self):
        from PyQt5.QtWidgets import QFormLayout

        main_layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.setStyleSheet("""
            QWidget {
                font-family: 'Segoe UI', sans-serif;
                font-size: 14px;
            }
            QLabel {
                color: #333;
                font-weight: bold;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 5px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 1px solid #007bff;
            }
            QPushButton {
                background-color: #007bff;
                color: white;
                border-radius: 5px;
                font-size: 14px;
                padding: 8px 15px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #004085;
            }
            QMessageBox {
                background-color: #f0f0f0;
                font-size: 14px;
            }
        """)

        # Helper to create file selection row
        def create_file_selection_row(label_text, line_edit_widget, browse_func):
            file_widget = QWidget()
            file_layout = QHBoxLayout(file_widget)
            file_layout.setContentsMargins(0, 0, 0, 0)
            line_edit_widget.setReadOnly(True)
            line_edit_widget.setStyleSheet("""
                QLineEdit {
                    background-color: #f5f5f5;
                }
            """)
            btn_browse = QPushButton("浏览...")
            btn_browse.clicked.connect(browse_func)
            btn_browse.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    border-radius: 5px;
                    font-size: 14px;
                    padding: 8px 15px;
                }
                QPushButton:hover {
                    background-color: #1976D2;
                }
                QPushButton:pressed {
                    background-color: #1565C0;
                }
            """)
            file_layout.addWidget(line_edit_widget)
            file_layout.addWidget(btn_browse)
            form_layout.addRow(label_text, file_widget)

        # 数据台账
        self.label_data = QLabel("未选择 数据台账 文件")
        self.data_file_display = QLineEdit(f"{os.path.join(self.data_file)}")
        create_file_selection_row("项目台账（读取）:", self.data_file_display, self.choose_data_file)

        # 写入模板
        self.label_template = QLabel("未选择 写入模板 文件")
        self.template_file_display = QLineEdit(f"{os.path.join(self.template_file)}")
        create_file_selection_row("验收测试报告（写入）:", self.template_file_display, self.choose_template_file)

        # 关键词
        self.input_keyword = QLineEdit()
        self.input_keyword.setPlaceholderText("请输入唯一任务书编号，例如：CPKFLX20230619001")
        form_layout.addRow("任务书编号:", self.input_keyword)

        # 额外字段
        extra_fields = [
            "测试单号", "申请理由", "开始时间", "结束时间", "测试依据", "测试范围"
        ]

        extra_tips =["CPKFLX20240426001","新产品导入","2025/07/14","2025/07/21","窗口式照相机技术需求规格书V1.3.xlsx","回归上一轮Bug、按照整机验收标准：底层软件、上层应用、机电安全"]

        for i, field in enumerate(extra_fields):
            input_box = QLineEdit()
            input_box.setPlaceholderText(f"可选填写 例如： {extra_tips[i]}")
            input_box.setStyleSheet("""
                QLineEdit {
                    background-color: #e9ecef;
                }
            """)
            form_layout.addRow(field + ":", input_box)
            self.input_fields[field] = input_box

        # 新增需求版本和测试版本数输入框
        self.input_fields["需求版本"] = QLineEdit()
        self.input_fields["需求版本"].setPlaceholderText("请输入需求版本，例如：1")
        self.input_fields["需求版本"].setStyleSheet("""
            QLineEdit {
                background-color: #fff3cd;
                border: 1px solid #ffeaa7;
            }
        """)
        form_layout.addRow("需求版本:", self.input_fields["需求版本"])

        self.input_fields["测试版本数"] = QLineEdit()
        self.input_fields["测试版本数"].setPlaceholderText("请输入测试版本数，例如：1")
        self.input_fields["测试版本数"].setStyleSheet("""
            QLineEdit {
                background-color: #fff3cd;
                border: 1px solid #ffeaa7;
            }
        """)
        form_layout.addRow("测试版本数:", self.input_fields["测试版本数"])

        main_layout.addLayout(form_layout)

        # 控制按钮区域（左右分布）
        button_layout = QHBoxLayout()
        btn_process = QPushButton("提取并写入")
        btn_process.clicked.connect(self.process)
        btn_process.setFixedHeight(40)
        btn_process.setStyleSheet("""
          QPushButton {
                background-color: #FF9800;
                color: white;
                border-radius: 5px;
                font-size: 16px;
                font-weight: bold;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background-color: #FB8C00;
            }
            QPushButton:pressed {
                background-color: #F57C00;
            }
        """)
        button_layout.addWidget(btn_process)

        btn_clear = QPushButton("清空所有输入")
        btn_clear.clicked.connect(self.clear_all_inputs)
        btn_clear.setStyleSheet("""
            QPushButton {
                background-color: #9E9E9E;
                color: white;
                border-radius: 8px;
                font-size: 16px;
                font-weight: bold;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background-color: #757575;
            }
            QPushButton:pressed {
                background-color: #616161;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
                color: #E0E0E0;
            }
        """)
        button_layout.addWidget(btn_clear)
        main_layout.addLayout(button_layout)

        # 日志窗口
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setStyleSheet("""
            QTextEdit {
                background-color: #f8f8f8;
                color: #333;
                font-family: 'Consolas', 'Monospace';
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 10px;
            }
        """)        
        main_layout.addWidget(self.log_output)

        self.setLayout(main_layout)

    def log(self, message: str, is_error: bool = False, clear_prev: bool = False):
        """带时间戳的日志输出，支持错误高亮和清空历史"""
        from PyQt5.QtCore import Qt
        from datetime import datetime
        if clear_prev:
            self.log_output.clear()
        cursor = self.log_output.textCursor()
        fmt = cursor.charFormat()
        if is_error:
            fmt.setForeground(Qt.red)
        else:
            fmt.setForeground(Qt.black)
        cursor.setCharFormat(fmt)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.log_output.append(f"[{timestamp}] {message}")
        self.log_output.ensureCursorVisible()

    def choose_data_file(self):
        initial_dir = os.path.dirname(self.data_file) if self.data_file else ""
        path, _ = QFileDialog.getOpenFileName(self, "选择 数据台账 文件", initial_dir, "Excel Files (*.xlsx *.xlsm)")
        if path:
            self.data_file = path
            self.data_file_display.setText(path)
            self.label_data.setText(f"已选择数据台账: {path}")
            # 保存到配置
            self.settings_manager.save_settings(
                "acceptance_filling",
                {"data_file": self.data_file, "template_file": self.template_file}
            )

    def choose_template_file(self):
        initial_dir = os.path.dirname(self.template_file) if self.template_file else ""
        path, _ = QFileDialog.getOpenFileName(self, "选择 写入模板 文件", initial_dir, "Excel Files (*.xlsx *.xlsm)")
        if path:
            self.template_file = path
            self.template_file_display.setText(path)
            self.label_template.setText(f"已选择模板文件: {path}")
            # 保存到配置
            self.settings_manager.save_settings(
                "acceptance_filling",
                {"data_file": self.data_file, "template_file": self.template_file}
            )

    def process(self):
        import traceback
        from core.excel_utils import fill_excel_template_acceptance
        if not self.data_file or not self.template_file:
            self.log("请先选择 数据台账 和 写入模板", is_error=True, clear_prev=False)
            QMessageBox.warning(self, "错误", "请先选择 数据台账 和 写入模板")
            return

        keyword = self.input_keyword.text().strip()
        if not keyword:
            self.log("请输入唯一的任务书编号", is_error=True, clear_prev=False)
            QMessageBox.warning(self, "错误", "请输入唯一的任务书编号")
            return

        try:
            self.log(f"开始处理任务书编号：{keyword}", clear_prev=False)
            self.log(f"数据台账文件: {self.data_file}", clear_prev=False)
            self.log(f"写入模板文件: {self.template_file}", clear_prev=False)
            for file_path, desc in [(self.data_file, '数据台账'), (self.template_file, '写入模板')]:
                if not os.path.isfile(file_path):
                    self.log(f"{desc}文件不存在: {file_path}", is_error=True, clear_prev=False)
                    raise FileNotFoundError(f"{desc}文件不存在: {file_path}")
                if not file_path.lower().endswith(('.xlsx', '.xlsm')):
                    self.log(f"{desc}文件不是Excel格式: {file_path}", is_error=True, clear_prev=False)
                    raise ValueError(f"{desc}文件不是Excel格式: {file_path}")
                size = os.path.getsize(file_path)
                self.log(f"{desc}文件大小: {size} 字节", clear_prev=False)
                if size == 0:
                    self.log(f"{desc}文件大小为0，文件内容异常: {file_path}", is_error=True, clear_prev=False)
                    raise ValueError(f"{desc}文件大小为0，文件内容异常: {file_path}")

            # 提取主数据
            target_fields = ['项目编号', '项目名称', '内部型号', '产品名称', '项目经理', '产品经理', '负责人']
            cell_mapping = {
                '项目编号': 'D2',
                '项目名称': 'H2',
                '项目经理': 'U2',
                '内部型号': 'D3',
                '产品名称': 'H3',
                '产品经理': 'U3',
                '负责人': 'U4'
            }

            result = find_row_by_fuzzy_column_value(
                file_path=self.data_file,
                key_column='任务书编号',
                key_value=keyword,
                target_columns=target_fields
            )

            if not result:
                self.log("台账中未找到匹配行", is_error=True, clear_prev=False)
                QMessageBox.warning(self, "未找到", "台账中未找到匹配行")
                return

            # 提取额外输入内容
            extra_data = {
                field: self.input_fields[field].text().strip()
                for field in self.input_fields
                if self.input_fields[field].text().strip()
            }

            extra_cell_mapping = {
                '测试单号': 'O2',
                '申请理由': 'D4',
                '开始时间': 'H4',
                '结束时间': 'O4',
                '测试依据': 'E6',
                '测试范围': 'E7'
            }

            # 合并数据
            all_data = result.copy()
            all_data.update(extra_data)
            
            # 创建字段映射配置
            field_mapping = {}
            
            # 主数据字段映射
            for field, cell in cell_mapping.items():
                field_mapping[field] = {"excel_cell": cell}
            
            # 额外字段映射
            for field, cell in extra_cell_mapping.items():
                field_mapping[field] = {"excel_cell": cell}

            # 使用新的智能填充函数
            self.log("准备使用智能填充功能写入数据到模板...", clear_prev=False)
            success = fill_excel_template_acceptance(
                template_path=self.template_file,
                data=all_data,
                field_mapping=field_mapping,
                sheet_name='验收测试结果',
                log_callback=self.log
            )
            if success:
                self.log("数据已成功写入 Excel 模板！", clear_prev=False)
                
                # 生成新的验收测试报告文档
                try:
                    self._generate_new_acceptance_report(all_data)
                except Exception as e:
                    self.log(f"生成新文档时出错: {str(e)}", is_error=True, clear_prev=False)
                
                QMessageBox.information(self, "成功", "数据已成功写入 Excel 模板！")
            else:
                self.log("写入失败，请检查日志。", is_error=True, clear_prev=False)
                QMessageBox.critical(self, "错误", "写入失败，请检查日志。")

        except Exception as e:
            tb = traceback.format_exc()
            self.log(f"发生错误：{type(e).__name__}: {str(e)}\n详细信息：\n{tb}", is_error=True, clear_prev=False)
            QMessageBox.critical(self, "错误", f"发生错误：\n{str(e)}")

    def clear_all_inputs(self):
        """清空所有输入框和日志内容"""
        for field in self.input_fields.values():
            field.clear()
        self.input_keyword.clear()
        self.data_file_display.clear()
        self.template_file_display.clear()
        self.log_output.clear()
        self.log("所有输入已清空。", clear_prev=True)

    def _generate_new_acceptance_report(self, data):
        """生成新的验收测试报告文档"""
        try:
            # 直接从传入的数据中获取项目信息，而不是从模板文件读取
            project_id = data.get('项目编号', '未知项目编号')
            project_name = data.get('项目名称', '未知项目名称')
            project_internal_name = data.get('内部型号', '未知型号')
            
            # 从模板文件名中提取版本号
            template_filename = os.path.basename(self.template_file)
            version_match = None
            
            # 尝试从文件名中提取版本号
            import re
            version_pattern = r'V(\d+\.\d+)'
            version_match = re.search(version_pattern, template_filename)
            
            if version_match:
                version = version_match.group(1)
                report_prefix = f"【验收测试报告V{version}】"
            else:
                # 如果无法从文件名提取，使用默认版本
                report_prefix = "【验收测试报告V3.2】"
                self.log(f"⚠️ 无法从模板文件名提取版本号，使用默认版本: {report_prefix}", clear_prev=False)
            
            # 获取当前日期
            current_date = datetime.now().strftime("%Y%m%d")
            
            # 构建新文件名
            new_filename = f"{report_prefix}{project_id}_{project_name}（{project_internal_name}）_{current_date}.xlsm"
            
            # 确保文件名中的特殊字符被处理
            new_filename = new_filename.replace('/', '_').replace('\\', '_').replace(':', '_')
            
            # 确定保存路径（保存到raw_data目录）
            raw_data_dir = "raw_data"
            if not os.path.exists(raw_data_dir):
                os.makedirs(raw_data_dir)
            
            new_file_path = os.path.join(raw_data_dir, new_filename)
            
            # 直接复制模板文件作为新文档
            shutil.copy2(self.template_file, new_file_path)
            
            self.log(f"✅ 新验收测试报告已生成: {new_filename}", clear_prev=False)
            self.log(f"📁 保存位置: {new_file_path}", clear_prev=False)

            # 在Sheet1写入标识“测试报告生成工具制作”
            try:
                from openpyxl import load_workbook
                is_xlsm = new_file_path.lower().endswith('.xlsm')
                wb = load_workbook(new_file_path, keep_vba=is_xlsm)
                ws = wb.worksheets[0]

                candidate_cells = ['Z1', 'AA1', 'AB1', 'AC1', 'A1']
                target_cell = 'A1'
                try:
                    for c in candidate_cells:
                        val = ws[c].value
                        if val is None or str(val).strip() == '':
                            target_cell = c
                            break
                except Exception:
                    target_cell = 'A1'

                ws[target_cell] = '测试报告生成工具制作'
                wb.save(new_file_path)
                self.log(f"已在Sheet1的 '{target_cell}' 写入标识：测试报告生成工具制作", clear_prev=False)
            except Exception as mark_err:
                self.log(f"⚠️ 写入Sheet1标识失败：{mark_err}", is_error=False, clear_prev=False)
            
        except Exception as e:
            self.log(f"❌ 生成新文档失败: {str(e)}", is_error=True, clear_prev=False)
            raise
