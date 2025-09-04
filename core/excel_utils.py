import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import xlwings as xw
import pandas as pd
import traceback




def find_row_by_fuzzy_column_value(file_path, key_column, key_value, target_columns):
    """
    在Excel文件中查找包含指定关键字的行，并返回目标列的值
    支持.xlsx和.xlsm文件，对于无法用openpyxl读取的.xlsm文件会使用xlwings作为备选方案
    """
    # 验证文件存在性和格式
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    if not file_path.lower().endswith(('.xlsx', '.xlsm')):
        raise ValueError(f"文件不是有效的Excel格式: {file_path}")
    
    # 检查文件大小
    file_size = os.path.getsize(file_path)
    if file_size == 0:
        raise ValueError(f"文件大小为0，文件可能损坏: {file_path}")
    
    # 对于.xlsm文件，优先尝试使用xlwings
    if file_path.lower().endswith('.xlsm'):
        try:
            return _find_row_with_xlwings(file_path, key_column, key_value, target_columns)
        except Exception as e:
            # 如果xlwings失败，尝试openpyxl
            pass
    
    # 使用openpyxl读取
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        if "File is not a zip file" in str(e):
            # 对于.xlsm文件，尝试使用xlwings作为备选方案
            if file_path.lower().endswith('.xlsm'):
                try:
                    return _find_row_with_xlwings(file_path, key_column, key_value, target_columns)
                except Exception as xlwings_error:
                    raise ValueError(f"无法读取.xlsm文件: {file_path}。openpyxl错误: {str(e)}。xlwings错误: {str(xlwings_error)}")
            else:
                raise ValueError(f"文件不是有效的Excel文件或已损坏: {file_path}。请确保文件是.xlsx或.xlsm格式且未损坏。")
        else:
            raise ValueError(f"无法打开Excel文件: {file_path}。错误: {str(e)}")
    
    try:
        sheet = wb.active
        
        # 获取表头
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {header: idx for idx, header in enumerate(headers) if header is not None}

        if key_column not in header_index:
            raise ValueError(f"找不到列标题: '{key_column}'。可用的列标题: {list(header_index.keys())}")
        
        for col in target_columns:
            if col not in header_index:
                raise ValueError(f"找不到目标列标题: '{col}'。可用的列标题: {list(header_index.keys())}")

        # 查找匹配的行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) > header_index[key_column]:
                cell_value = str(row[header_index[key_column]]) if row[header_index[key_column]] is not None else ""
                if key_value in cell_value:
                    return {col: row[header_index[col]] for col in target_columns}
        
        return None
        
    except Exception as e:
        raise ValueError(f"处理Excel文件时发生错误: {str(e)}")
    finally:
        try:
            wb.close()
        except:
            pass


def _find_row_with_xlwings(file_path, key_column, key_value, target_columns):
    """
    使用xlwings读取Excel文件并查找匹配的行
    """
    app = None
    wb = None
    try:
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        sheet = wb.sheets[0]  # 使用第一个工作表
        
        # 获取表头
        headers = sheet.range('A1').expand('right').value
        if not headers:
            raise ValueError("无法读取表头")
        
        # 创建列索引映射
        header_index = {str(header): idx for idx, header in enumerate(headers) if header is not None}
        
        if key_column not in header_index:
            raise ValueError(f"找不到列标题: '{key_column}'。可用的列标题: {list(header_index.keys())}")
        
        for col in target_columns:
            if col not in header_index:
                raise ValueError(f"找不到目标列标题: '{col}'。可用的列标题: {list(header_index.keys())}")
        
        # 获取所有数据
        data_range = sheet.range('A1').expand('table')
        all_data = data_range.value
        
        if not all_data or len(all_data) < 2:
            return None
        
        # 查找匹配的行
        for row_idx, row in enumerate(all_data[1:], 2):  # 从第二行开始
            if len(row) > header_index[key_column]:
                cell_value = str(row[header_index[key_column]]) if row[header_index[key_column]] is not None else ""
                if key_value in cell_value:
                    return {col: row[header_index[col]] for col in target_columns}
        
        return None
        
    except Exception as e:
        raise ValueError(f"使用xlwings读取Excel文件时发生错误: {str(e)}")
    finally:
        try:
            if wb:
                wb.close()
            if app:
                app.quit()
        except:
            pass


def write_to_target_sheet(file_path, sheet_name, cell_map, data_dict):
    """
    将数据写入Excel工作表的指定单元格
    支持.xlsx和.xlsm文件，对于无法用openpyxl写入的.xlsm文件会使用xlwings作为备选方案
    """
    # 验证文件存在性和格式
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    if not file_path.lower().endswith(('.xlsx', '.xlsm')):
        raise ValueError(f"文件不是有效的Excel格式: {file_path}")
    
    # 检查文件大小
    file_size = os.path.getsize(file_path)
    if file_size == 0:
        raise ValueError(f"文件大小为0，文件可能损坏: {file_path}")
    
    # 对于.xlsm文件，优先尝试使用xlwings
    if file_path.lower().endswith('.xlsm'):
        try:
            return _write_with_xlwings(file_path, sheet_name, cell_map, data_dict)
        except Exception as e:
            # 如果xlwings失败，尝试openpyxl
            pass
    
    # 使用openpyxl写入
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        if "File is not a zip file" in str(e):
            # 对于.xlsm文件，尝试使用xlwings作为备选方案
            if file_path.lower().endswith('.xlsm'):
                try:
                    return _write_with_xlwings(file_path, sheet_name, cell_map, data_dict)
                except Exception as xlwings_error:
                    raise ValueError(f"无法写入.xlsm文件: {file_path}。openpyxl错误: {str(e)}。xlwings错误: {str(xlwings_error)}")
            else:
                raise ValueError(f"文件不是有效的Excel文件或已损坏: {file_path}。请确保文件是.xlsx或.xlsm格式且未损坏。")
        else:
            raise ValueError(f"无法打开Excel文件: {file_path}。错误: {str(e)}")
    
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"找不到工作表：{sheet_name}。可用的工作表: {wb.sheetnames}")
        
        sheet = wb[sheet_name]
        for key, cell in cell_map.items():
            sheet[cell] = data_dict.get(key, "")
        
        wb.save(file_path)
        
    except Exception as e:
        raise ValueError(f"写入Excel文件时发生错误: {str(e)}")
    finally:
        try:
            wb.close()
        except:
            pass


def _write_with_xlwings(file_path, sheet_name, cell_map, data_dict):
    """
    使用xlwings写入Excel文件
    """
    app = None
    wb = None
    try:
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        
        # 检查工作表是否存在
        sheet_names = [sheet.name for sheet in wb.sheets]
        if sheet_name not in sheet_names:
            raise ValueError(f"找不到工作表：{sheet_name}。可用的工作表: {sheet_names}")
        
        sheet = wb.sheets[sheet_name]
        
        # 写入数据
        for key, cell in cell_map.items():
            value = data_dict.get(key, "")
            sheet.range(cell).value = value
        
        wb.save()
        return True
        
    except Exception as e:
        raise ValueError(f"使用xlwings写入Excel文件时发生错误: {str(e)}")
    finally:
        try:
            if wb:
                wb.close()
            if app:
                app.quit()
        except:
            pass



def fill_excel_template_acceptance(template_path: str, data: dict, field_mapping: dict, sheet_name: str, log_callback=None):
    """
    智能填充验收测试结果Excel模板
    支持新模板和上一轮测试报告的智能识别和处理
    """
    if not os.path.exists(template_path):
        if log_callback: log_callback(f"错误: Excel 模板文件未找到于 '{template_path}'", is_error=True)
        return False
    if not template_path.lower().endswith((".xlsx", ".xlsm")):
        if log_callback: log_callback(f"错误: 提供的文件 '{template_path}' 不是有效的 Excel 模板 (.xlsx 或 .xlsm)。", is_error=True)
        return False

    # 对于.xlsm文件，使用xlwings处理
    if template_path.lower().endswith('.xlsm'):
        return _fill_excel_template_with_xlwings(template_path, data, field_mapping, sheet_name, log_callback)
    
    # 对于.xlsx文件，使用openpyxl处理
    return _fill_excel_template_with_openpyxl(template_path, data, field_mapping, sheet_name, log_callback)


def _fill_excel_template_with_xlwings(template_path: str, data: dict, field_mapping: dict, sheet_name: str, log_callback=None):
    """
    使用xlwings填充Excel模板（适用于.xlsm文件）
    """
    app = None
    wb = None
    try:
        app = xw.App(visible=False)
        wb = app.books.open(template_path)
        
        if sheet_name not in [sheet.name for sheet in wb.sheets]:
            if log_callback: log_callback(f"错误: Excel 工作簿中未找到名为 '{sheet_name}' 的工作表。", is_error=True)
            return False
        
        ws = wb.sheets[sheet_name]
        if log_callback: log_callback(f"已成功加载工作表: '{sheet_name}'。")
        
        # 检查第18行是否有内容来判断是新模板还是上一轮测试报告
        # 检查B18和C18单元格，因为数据通常从B列开始
        row_18_content_b = ws.range('B18').value
        row_18_content_c = ws.range('C18').value
        is_new_template = (row_18_content_b is None or str(row_18_content_b).strip() == "") and \
                         (row_18_content_c is None or str(row_18_content_c).strip() == "")
        
        if log_callback: 
            if is_new_template:
                log_callback("检测到新模板：第18行无内容，将使用B17和D17位置填写需求版本和测试版本数", is_error=False)
            else:
                log_callback("检测到上一轮测试报告：第18行有内容，将寻找空白行并复制下拉框公式", is_error=False)
        
        # 处理需求版本和测试版本数
        new_row = None
        if "需求版本" in data and data["需求版本"]:
            if is_new_template:
                # 新模板：B17位置
                ws.range('B17').value = data["需求版本"]
                if log_callback: log_callback(f"  写入需求版本: '{data['需求版本']}' 到单元格 'B17'")
            else:
                # 上一轮测试报告：先处理下拉框公式复制，然后填入数据
                new_row = _handle_dropdown_formula_copy(ws, log_callback)
                if new_row:
                    ws.range(f'B{new_row}').value = data["需求版本"]
                    if log_callback: log_callback(f"  写入需求版本: '{data['需求版本']}' 到单元格 'B{new_row}'")
        
        if "测试版本数" in data and data["测试版本数"]:
            if is_new_template:
                # 新模板：D17位置
                ws.range('D17').value = data["测试版本数"]
                if log_callback: log_callback(f"  写入测试版本数: '{data['测试版本数']}' 到单元格 'D17'")
            else:
                # 上一轮测试报告：使用之前创建的新行
                if new_row is None:
                    new_row = _handle_dropdown_formula_copy(ws, log_callback)
                if new_row:
                    ws.range(f'D{new_row}').value = data["测试版本数"]
                    if log_callback: log_callback(f"  写入测试版本数: '{data['测试版本数']}' 到单元格 'D{new_row}'")
        
        # 处理其他字段
        for field_name, config in field_mapping.items():
            if field_name in ["需求版本", "测试版本数"]:
                continue  # 已经处理过了
            
            excel_cell_coord = config.get("excel_cell", "")
            value = data.get(field_name, "")
            
            if excel_cell_coord and value:
                try:
                    ws.range(excel_cell_coord).value = value
                    if log_callback: log_callback(f"  写入字段 '{field_name}': '{value}' 到单元格 '{excel_cell_coord}'")
                except Exception as e:
                    if log_callback: log_callback(f"  写入字段 '{field_name}' 到单元格 '{excel_cell_coord}' 失败。原因: {e}", is_error=True)
        
        # 下拉框公式复制已在处理需求版本和测试版本数时完成
        
        # 保存文件到原文档
        wb.save(template_path)
        
        if log_callback: log_callback(f"\n--- 成功填充！文件已保存到原文档: '{template_path}' ---", is_error=False)
        return True
        
    except Exception as e:
        if log_callback: log_callback(f"错误: 使用xlwings处理Excel文件时发生错误: {e}", is_error=True)
        return False
    finally:
        try:
            if wb:
                wb.close()
            if app:
                app.quit()
        except:
            pass


def _fill_excel_template_with_openpyxl(template_path: str, data: dict, field_mapping: dict, sheet_name: str, log_callback=None):
    """
    使用openpyxl填充Excel模板（适用于.xlsx文件）
    """
    try:
        wb = load_workbook(template_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if log_callback: log_callback(f"已成功加载工作表: '{sheet_name}'。")
        else:
            if log_callback: log_callback(f"错误: Excel 工作簿中未找到名为 '{sheet_name}' 的工作表。可用的工作表: {wb.sheetnames}", is_error=True)
            return False
    except Exception as e:
        if "File is not a zip file" in str(e):
            if log_callback: log_callback(f"错误: 文件不是有效的Excel文件或已损坏: '{template_path}'。请确保文件是.xlsx或.xlsm格式且未损坏。", is_error=True)
        else:
            if log_callback: log_callback(f"错误: 无法加载 Excel 工作簿或获取指定工作表 '{template_path}'。原因: {e}", is_error=True)
        return False

    # 处理合并单元格映射
    merged_cells_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_coord = ws.cell(row=min_row, column=min_col).coordinate
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = ws.cell(row=row, column=col).coordinate
                merged_cells_map[coord] = top_left_coord

    # 检查第18行是否有内容来判断是新模板还是上一轮测试报告
    # 检查B18和C18单元格，因为数据通常从B列开始
    row_18_content_b = ws['B18'].value
    row_18_content_c = ws['C18'].value
    is_new_template = (row_18_content_b is None or str(row_18_content_b).strip() == "") and \
                     (row_18_content_c is None or str(row_18_content_c).strip() == "")
    
    if log_callback: 
        if is_new_template:
            log_callback("检测到新模板：第18行无内容，将使用B17和D17位置填写需求版本和测试版本数", is_error=False)
        else:
            log_callback("检测到上一轮测试报告：第18行有内容，将寻找空白行并复制下拉框公式", is_error=False)

    if log_callback: log_callback("正在写入数据到 Excel...")
    all_fields_processed_successfully = True

    # 处理需求版本和测试版本数
    new_row = None
    if "需求版本" in data and data["需求版本"]:
        if is_new_template:
            # 新模板：B17位置
            actual_cell_coord = merged_cells_map.get('B17', 'B17')
            try:
                ws[actual_cell_coord] = data["需求版本"]
                if log_callback: log_callback(f"  写入需求版本: '{data['需求版本']}' 到单元格 '{actual_cell_coord}'")
            except Exception as e:
                if log_callback: log_callback(f"  写入需求版本失败。原因: {e}", is_error=True)
                all_fields_processed_successfully = False
        else:
            # 上一轮测试报告：先处理下拉框公式复制，然后填入数据
            new_row = _handle_dropdown_formula_copy_openpyxl(ws, log_callback)
            if new_row:
                try:
                    ws[f'B{new_row}'] = data["需求版本"]
                    if log_callback: log_callback(f"  写入需求版本: '{data['需求版本']}' 到单元格 'B{new_row}'")
                except Exception as e:
                    if log_callback: log_callback(f"  写入需求版本失败。原因: {e}", is_error=True)
                    all_fields_processed_successfully = False

    if "测试版本数" in data and data["测试版本数"]:
        if is_new_template:
            # 新模板：D17位置
            actual_cell_coord = merged_cells_map.get('D17', 'D17')
            try:
                ws[actual_cell_coord] = data["测试版本数"]
                if log_callback: log_callback(f"  写入测试版本数: '{data['测试版本数']}' 到单元格 '{actual_cell_coord}'")
            except Exception as e:
                if log_callback: log_callback(f"  写入测试版本数失败。原因: {e}", is_error=True)
                all_fields_processed_successfully = False
        else:
            # 上一轮测试报告：使用之前创建的新行
            if new_row is None:
                new_row = _handle_dropdown_formula_copy_openpyxl(ws, log_callback)
            if new_row:
                try:
                    ws[f'D{new_row}'] = data["测试版本数"]
                    if log_callback: log_callback(f"  写入测试版本数: '{data['测试版本数']}' 到单元格 'D{new_row}'")
                except Exception as e:
                    if log_callback: log_callback(f"  写入测试版本数失败。原因: {e}", is_error=True)
                    all_fields_processed_successfully = False

    # 处理其他字段
    for field_name, config in field_mapping.items():
        if field_name in ["需求版本", "测试版本数"]:
            continue  # 已经处理过了
        
        excel_cell_coord = config.get("excel_cell", "")
        value = data.get(field_name, "")

        if excel_cell_coord and value:
            actual_cell_coord = merged_cells_map.get(excel_cell_coord, excel_cell_coord)
            try:
                ws[actual_cell_coord] = value
                if log_callback: log_callback(f"  写入字段 '{field_name}': '{value}' 到单元格 '{actual_cell_coord}'")
            except Exception as e:
                if log_callback: log_callback(f"  写入字段 '{field_name}' 到单元格 '{actual_cell_coord}' 失败。原因: {e}", is_error=True)
                all_fields_processed_successfully = False

    try:
        wb.save(template_path)
        if all_fields_processed_successfully:
            if log_callback: log_callback(f"\n--- 成功填充！文件已保存到原文档: '{template_path}' ---", is_error=False)
        else:
            if log_callback: log_callback(f"\n--- 填充完成，但有部分字段出现问题。文件已保存到原文档: '{template_path}' ---", is_error=False)
            if log_callback: log_callback("注意: 请检查日志，有部分字段未填写内容或写入失败。", is_error=True)
        return True
    except PermissionError:
        if log_callback: log_callback(f"错误: 无法保存文件 '{template_path}'。原因: 权限被拒绝，请确保 Excel 文件已关闭且您有写入权限。", is_error=True)
        return False
    except Exception as e:
        if log_callback: log_callback(f"错误: 无法将填充后的 Excel 文件保存到 '{template_path}'。原因: {e}", is_error=True)
        return False


def _find_blank_row_for_version_info(ws, log_callback=None):
    """
    在上一轮测试报告中寻找适合填写版本信息的空白行
    """
    try:
        # 从第18行开始向下查找空白行
        for row in range(18, 50):  # 限制在50行内查找
            cell_value_b = ws.range(f'B{row}').value
            cell_value_c = ws.range(f'C{row}').value
            if (cell_value_b is None or str(cell_value_b).strip() == "") and \
               (cell_value_c is None or str(cell_value_c).strip() == ""):
                if log_callback: log_callback(f"  找到空白行: 第{row}行", is_error=False)
                return row
        
        if log_callback: log_callback("  警告: 未找到合适的空白行，使用第18行", is_error=True)
        return 18
    except Exception as e:
        if log_callback: log_callback(f"  查找空白行时出错: {e}", is_error=True)
        return 18


def _find_blank_row_for_version_info_openpyxl(ws, log_callback=None):
    """
    在上一轮测试报告中寻找适合填写版本信息的空白行（openpyxl版本）
    """
    try:
        # 从第18行开始向下查找空白行
        for row in range(18, 50):  # 限制在50行内查找
            cell_value_b = ws[f'B{row}'].value
            cell_value_c = ws[f'C{row}'].value
            if (cell_value_b is None or str(cell_value_b).strip() == "") and \
               (cell_value_c is None or str(cell_value_c).strip() == ""):
                if log_callback: log_callback(f"  找到空白行: 第{row}行", is_error=False)
                return row
        
        if log_callback: log_callback("  警告: 未找到合适的空白行，使用第18行", is_error=True)
        return 18
    except Exception as e:
        if log_callback: log_callback(f"  查找空白行时出错: {e}", is_error=True)
        return 18


def _handle_dropdown_formula_copy(ws, log_callback=None):
    """
    处理下拉框公式复制（仅适用于上一轮测试报告）
    1. 找到有内容的最后一行
    2. 将最后一行的B到W列复制到下一行空白行（保留公式）
    3. 将之前所有有内容的行的B到W列转换为纯数值（移除公式）
    4. 在新复制的行中填入用户输入的需求版本和测试版本数
    """
    try:
        if log_callback: log_callback("  开始处理下拉框公式复制...", is_error=False)
        
        # 1. 找到有内容的最后一行
        last_data_row = None
        for row in range(17, 50):  # 从17行开始查找
            cell_value_b = ws.range(f'B{row}').value
            cell_value_c = ws.range(f'C{row}').value
            if (cell_value_b is not None and str(cell_value_b).strip() != "") or \
               (cell_value_c is not None and str(cell_value_c).strip() != ""):
                last_data_row = row
            else:
                # 如果当前行为空，且已经找到了有内容的行，则停止查找
                if last_data_row is not None:
                    break
        
        if last_data_row is None:
            if log_callback: log_callback("  警告: 未找到有内容的行", is_error=True)
            return
        
        if log_callback: log_callback(f"  找到有内容的最后一行: 第{last_data_row}行", is_error=False)
        
        # 2. 找到下一个空白行
        blank_row = last_data_row + 1
        if log_callback: log_callback(f"  将在第{blank_row}行创建新数据", is_error=False)
        
        # 3. 将最后一行的B到W列复制到下一行空白行（保留公式）
        try:
            # 复制B到W列的内容和公式
            source_range = ws.range(f'B{last_data_row}:W{last_data_row}')
            target_range = ws.range(f'B{blank_row}:W{blank_row}')
            
            # 复制公式
            source_range.api.Copy(target_range.api)
            if log_callback: log_callback(f"  已复制第{last_data_row}行的B到W列到第{blank_row}行", is_error=False)
            
        except Exception as e:
            if log_callback: log_callback(f"  复制行数据时出错: {e}", is_error=True)
        
        # 4. 将之前所有有内容的行的B到W列转换为纯数值（移除公式）
        try:
            for row in range(17, last_data_row + 1):
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
                    try:
                        cell = ws.range(f'{col}{row}')
                        if cell.formula and cell.formula.startswith('='):
                            # 保存当前值
                            current_value = cell.value
                            # 清除公式，只保留数值
                            cell.formula = None
                            cell.value = current_value
                    except:
                        continue  # 跳过无效的单元格
            
            if log_callback: log_callback(f"  已移除第17行到第{last_data_row}行的公式，只保留数值", is_error=False)
        except Exception as e:
            if log_callback: log_callback(f"  移除公式时出错: {e}", is_error=True)
        
        # 5. 返回新创建的行号，供后续填入需求版本和测试版本数
        return blank_row
        
    except Exception as e:
        if log_callback: log_callback(f"  处理下拉框公式复制时出错: {e}", is_error=True)
        return None


def _handle_dropdown_formula_copy_openpyxl(ws, log_callback=None):
    """
    处理下拉框公式复制（仅适用于上一轮测试报告）- openpyxl版本
    1. 找到有内容的最后一行
    2. 将最后一行的B到W列复制到下一行空白行（保留公式）
    3. 将之前所有有内容的行的B到W列转换为纯数值（移除公式）
    4. 在新复制的行中填入用户输入的需求版本和测试版本数
    """
    try:
        if log_callback: log_callback("  开始处理下拉框公式复制...", is_error=False)
        
        # 1. 找到有内容的最后一行
        last_data_row = None
        for row in range(17, 50):  # 从17行开始查找
            cell_value_b = ws[f'B{row}'].value
            cell_value_c = ws[f'C{row}'].value
            if (cell_value_b is not None and str(cell_value_b).strip() != "") or \
               (cell_value_c is not None and str(cell_value_c).strip() != ""):
                last_data_row = row
            else:
                # 如果当前行为空，且已经找到了有内容的行，则停止查找
                if last_data_row is not None:
                    break
        
        if last_data_row is None:
            if log_callback: log_callback("  警告: 未找到有内容的行", is_error=True)
            return None
        
        if log_callback: log_callback(f"  找到有内容的最后一行: 第{last_data_row}行", is_error=False)
        
        # 2. 找到下一个空白行
        blank_row = last_data_row + 1
        if log_callback: log_callback(f"  将在第{blank_row}行创建新数据", is_error=False)
        
        # 3. 将最后一行的B到W列复制到下一行空白行（保留公式）
        try:
            # 复制B到W列的内容和公式
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
                source_cell = ws[f'{col}{last_data_row}']
                target_cell = ws[f'{col}{blank_row}']
                target_cell.value = source_cell.value
                # 注意：openpyxl对公式的支持有限，这里主要复制值
            
            if log_callback: log_callback(f"  已复制第{last_data_row}行的B到W列到第{blank_row}行", is_error=False)
            
        except Exception as e:
            if log_callback: log_callback(f"  复制行数据时出错: {e}", is_error=True)
        
        # 4. 将之前所有有内容的行的B到W列转换为纯数值（移除公式）
        try:
            for row in range(17, last_data_row + 1):
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
                    try:
                        cell = ws[f'{col}{row}']
                        if cell.value and str(cell.value).startswith('='):
                            # 保存当前值
                            current_value = cell.value
                            # 清除公式，只保留数值
                            cell.value = current_value
                    except:
                        continue  # 跳过无效的单元格
            
            if log_callback: log_callback(f"  已移除第17行到第{last_data_row}行的公式，只保留数值", is_error=False)
        except Exception as e:
            if log_callback: log_callback(f"  移除公式时出错: {e}", is_error=True)
        
        # 5. 返回新创建的行号，供后续填入需求版本和测试版本数
        return blank_row
        
    except Exception as e:
        if log_callback: log_callback(f"  处理下拉框公式复制时出错: {e}", is_error=True)
        return None


def consolidate_excel_data_and_insert_chart(doc1_path: str, doc2_path: str, doc3_path: str, doc4_path: str,
                                            target_report_path: str, 
                                            append_mode_doc1: bool = False,
                                            append_mode_doc2: bool = False, 
                                            append_mode_doc3: bool = False,
                                            log_callback=None):
    """
    Core data consolidation logic using xlwings: copies three data tables (starting from the second row)
    to the third row of corresponding sheets in the target report, preserving format and sheet order.
    Also inserts the device appearance image into the '设备外观图' sheet.
    This version allows individual source documents (Doc1-Doc4) to be optional.
    支持追加模式：当append_mode为True时，数据将追加到现有数据后面而不是覆盖。
    """
    source_info = [
        {'path': doc1_path, 'sheet_name': '遗留缺陷列表', 'append_mode': append_mode_doc1},
        {'path': doc2_path, 'sheet_name': '产品需求列表', 'append_mode': append_mode_doc2},
        {'path': doc3_path, 'sheet_name': '验收测试用例', 'append_mode': append_mode_doc3}
    ]

    app = None
    try:
        # Check target report path first as it's mandatory
        if not target_report_path or not os.path.exists(target_report_path):
            if log_callback: log_callback(
                f"错误：目标报告文件 '{os.path.basename(target_report_path) if target_report_path else '未指定'}' 不存在或路径为空。",
                True)
            return False

        # 清理COM组件缓存，解决打包后的null bytes问题
        try:
            import win32com.client
            import pythoncom
            # 清理COM组件缓存
            pythoncom.CoInitialize()
            if log_callback: log_callback("COM组件初始化完成", False)
        except Exception as com_error:
            if log_callback: log_callback(f"COM组件初始化警告: {com_error}", False)

        # 尝试初始化xlwings，如果失败则使用备用方案
        try:
            app = xw.App(visible=False, add_book=False)
            if log_callback: log_callback("xlwings应用程序初始化成功", False)
        except Exception as xlwings_error:
            if log_callback: log_callback(f"xlwings初始化失败，尝试备用方案: {xlwings_error}", True)

          # 如果xlwings失败，尝试使用win32com作为备用方案
            return _consolidate_with_win32com_fallback(doc1_path, doc2_path, doc3_path, doc4_path, 
                                                    target_report_path, append_mode_doc1, append_mode_doc2, 
                                                    append_mode_doc3, log_callback)
        wb = app.books.open(target_report_path, update_links=False)
        if log_callback: log_callback(f"已打开目标报告：{os.path.basename(target_report_path)}", False)

        for info in source_info:
            src_path = info['path']
            target_sheet_name = info['sheet_name']
            append_mode = info['append_mode']

            if not src_path or not os.path.exists(src_path):
                if log_callback: log_callback(
                    f"警告：源文件 '{os.path.basename(src_path) if src_path else target_sheet_name + '文档'}' 未选择或不存在，跳过处理。",
                    False)
                continue  # Skip to the next source file

            mode_text = "追加" if append_mode else "覆盖"
            if log_callback: log_callback(
                f"\n正在处理源文件 '{os.path.basename(src_path)}' -> 工作表 '{target_sheet_name}' [模式: {mode_text}]", False)

            try:
                df = pd.read_excel(src_path, header=None)
                if df.empty:
                    if log_callback: log_callback(f"警告：源文件 '{os.path.basename(src_path)}' 为空或无数据。", False)
                    data = []
                else:
                    data = df.iloc[1:].values.tolist()  # Skip the first row (header in ZenTao exports)
                    
                    # 模块2：对于"遗留缺陷列表"，如果第10列出现"已关闭"则不新增且删除此行
                    if target_sheet_name == '遗留缺陷列表' and data:
                        original_count = len(data)
                        
                        # 第10列的索引是9（从0开始计数）
                        bug_status_col_index = 9
                        
                        if log_callback:
                            log_callback(f"🔍 遗留缺陷列表：检查第10列（索引{bug_status_col_index}）的'已关闭'状态", False)
                        
                        # 过滤掉第10列为"已关闭"的行
                        filtered_data = []
                        removed_count = 0
                        
                        for row_index, row in enumerate(data):
                            if len(row) > bug_status_col_index:
                                bug_status = str(row[bug_status_col_index]).strip() if row[bug_status_col_index] else ""
                                bug_id = str(row[0]).strip() if len(row) > 0 and row[0] else "未知"
                                
                                if bug_status != "已关闭":
                                    filtered_data.append(row)
                                else:
                                    removed_count += 1
                                    if log_callback:
                                        log_callback(f"🗑️ 删除第{row_index+1}行: Bug编号={bug_id}, 第10列状态='{bug_status}'", False)
                            else:
                                # 如果行数据不完整（少于10列），保留该行
                                filtered_data.append(row)
                                if log_callback:
                                    log_callback(f"⚠️ 保留第{row_index+1}行: 数据不完整（少于10列）", False)
                        
                        data = filtered_data
                        
                        if log_callback:
                            log_callback(f"✅ 遗留缺陷列表过滤完成：原始 {original_count} 行，删除 {removed_count} 行'已关闭'状态，保留 {len(data)} 行", False)
                                
            except Exception as e:
                if log_callback: log_callback(f"错误: 读取源文件 '{os.path.basename(src_path)}' 失败。原因: {e}", True)
                if log_callback: log_callback(traceback.format_exc(), True)
                continue  # Skip to the next source file

            if target_sheet_name in [s.name for s in wb.sheets]:
                sht = wb.sheets[target_sheet_name]
                if log_callback: log_callback(f"已找到工作表: '{target_sheet_name}'", False)
            else:
                try:
                    sht = wb.sheets.add(name=target_sheet_name, after=wb.sheets[-1])  # Add new sheet at the end
                    if log_callback: log_callback(f"创建工作表：{target_sheet_name}", False)
                except Exception as e:
                    if log_callback: log_callback(f"错误: 无法创建工作表 '{target_sheet_name}'。原因: {e}", True)
                    if log_callback: log_callback(traceback.format_exc(), True)
                    sht = None  # Ensure sht is None if creation failed

            if sht:  # Only proceed if sheet exists or was created
                start_row_excel = 3
                used_range = sht.used_range
                
                if append_mode:
                    # 追加模式：找到真正有数据内容的最后一行，从下一行开始插入
                    if used_range and used_range.api is not None:
                        # 从第3行开始向下扫描，找到最后一行有实际数据的行
                        last_data_row = start_row_excel - 1  # 默认从第3行开始
                        
                        # 扫描从第3行到used_range最后一行，找到真正有数据的最后一行
                        max_scan_row = used_range.last_cell.row
                        for row_num in range(start_row_excel, max_scan_row + 1):
                            try:
                                # 检查这一行是否有非空的数据（排除空格和None）
                                row_values = sht.range(f'{row_num}:{row_num}').value
                                if isinstance(row_values, list):
                                    # 检查是否有非空非None的值
                                    has_data = any(cell is not None and str(cell).strip() != '' for cell in row_values if cell is not None)
                                else:
                                    # 单个值的情况
                                    has_data = row_values is not None and str(row_values).strip() != ''
                                
                                if has_data:
                                    last_data_row = row_num
                            except Exception:
                                # 如果读取行出错，继续下一行
                                continue
                        
                        start_row_excel = last_data_row + 1
                        if log_callback: log_callback(
                            f"追加模式: 扫描到第 {last_data_row} 行为最后有数据的行，将从第 {start_row_excel} 行开始插入数据（原有数据保留）", False)
                    else:
                        if log_callback: log_callback(
                            f"追加模式: 工作表为空，从第 {start_row_excel} 行开始插入", False)
                    
                    # 追加模式下不清除原有数据
                else:
                    # 覆盖模式：清除原有数据
                    last_row_to_clear = used_range.last_cell.row if used_range and used_range.api is not None else start_row_excel
                    last_col_to_clear = used_range.last_cell.column if used_range and used_range.api is not None else (
                        df.shape[1] if not df.empty else 1)

                    if data:
                        max_cols_in_data = max(len(row) for row in data)
                        last_col_to_clear = max(last_col_to_clear, max_cols_in_data)

                    if last_row_to_clear >= start_row_excel:
                        if log_callback: log_callback(
                            f"覆盖模式: 清除 '{target_sheet_name}' 第 {start_row_excel} 行到第 {last_row_to_clear} 行的内容 (到第 {last_col_to_clear} 列)...",
                            False)
                        try:
                            sht.range((start_row_excel, 1), (last_row_to_clear, last_col_to_clear)).clear_contents()
                        except Exception as e:
                            if log_callback: log_callback(f"错误: 清除工作表 '{target_sheet_name}' 旧数据失败。原因: {e}",
                                                          True)
                            if log_callback: log_callback(traceback.format_exc(), True)
                            continue  # Try to proceed, but log error
                    else:
                        if log_callback: log_callback(f"覆盖模式: 工作表 '{target_sheet_name}' 已经足够干净，无需清除旧数据。", False)

                if data:
                    if log_callback: log_callback(f"粘贴新数据到 '{target_sheet_name}' 第 {start_row_excel} 行，共 {len(data)} 行。", False)
                    try:
                        sht.range((start_row_excel, 1)).value = data
                        
                        # 复制格式功能：将现有数据行的格式应用到新添加的数据行（性能优化版本）
                        if log_callback: log_callback(f"正在复制格式到新添加的数据行...", False)
                        
                        # 找到可以作为模板的现有数据行
                        template_row = None
                        if start_row_excel > 3:  # 如果是追加模式且有现有数据
                            # 使用前一行作为模板
                            template_row = start_row_excel - 1
                        else:
                            # 如果是覆盖模式或无现有数据，尝试使用第3行作为模板（如果存在的话）
                            try:
                                test_row = sht.range('3:3').value
                                if test_row and any(cell is not None for cell in (test_row if isinstance(test_row, list) else [test_row])):
                                    template_row = 3
                            except Exception:
                                pass
                        
                        if template_row and len(data) > 0:
                            try:
                                # 获取数据的列数
                                data_cols = len(data[0]) if data and len(data[0]) > 0 else 1
                                end_row_excel = start_row_excel + len(data) - 1
                                
                                # 性能优化：使用批量格式复制
                                if log_callback: log_callback(
                                    f"批量复制格式：模板行{template_row} -> 目标行{start_row_excel}-{end_row_excel}（{len(data)}行）", False)
                                
                                try:
                                    # 方法1：批量复制整个区域（最快）
                                    template_range = sht.range(f'{template_row}:{template_row}')
                                    target_range = sht.range(f'{start_row_excel}:{end_row_excel}')
                                    
                                    # 一次性复制模板行格式到所有目标行
                                    template_range.api.Copy()
                                    target_range.api.PasteSpecial(Paste=-4122)  # xlPasteFormats
                                    
                                    if log_callback: log_callback(
                                        f"✅ 批量格式复制成功：{len(data)}行格式已应用", False)
                                        
                                except Exception as batch_error:
                                    # 方法2：如果批量复制失败，使用分批复制（中等速度）
                                    if log_callback: log_callback(
                                        f"批量复制失败，尝试分批复制。原因: {batch_error}", False)
                                    
                                    try:
                                        # 每10行为一批进行复制
                                        batch_size = 10
                                        successful_batches = 0
                                        
                                        for i in range(0, len(data), batch_size):
                                            batch_start = start_row_excel + i
                                            batch_end = min(start_row_excel + i + batch_size - 1, end_row_excel)
                                            
                                            try:
                                                template_range = sht.range(f'{template_row}:{template_row}')
                                                batch_target_range = sht.range(f'{batch_start}:{batch_end}')
                                                
                                                template_range.api.Copy()
                                                batch_target_range.api.PasteSpecial(Paste=-4122)
                                                successful_batches += 1
                                                
                                            except Exception:
                                                # 如果某个批次失败，继续下一个批次
                                                continue
                                        
                                        if log_callback: log_callback(
                                            f"✅ 分批格式复制完成：{successful_batches}个批次成功", False)
                                            
                                    except Exception as fallback_error:
                                        # 方法3：如果分批也失败，只对前几行进行格式复制（保底方案）
                                        if log_callback: log_callback(
                                            f"分批复制也失败，使用保底方案。原因: {fallback_error}", False)
                                        
                                        try:
                                            # 只对前5行进行格式复制，避免长时间等待
                                            limited_rows = min(5, len(data))
                                            template_range = sht.range(f'{template_row}:{template_row}')
                                            
                                            for i in range(limited_rows):
                                                target_row = start_row_excel + i
                                                target_range = sht.range(f'{target_row}:{target_row}')
                                                
                                                template_range.api.Copy()
                                                target_range.api.PasteSpecial(Paste=-4122)
                                            
                                            if log_callback: log_callback(
                                                f"⚠️ 保底格式复制完成：仅前{limited_rows}行应用了格式", False)
                                                
                                        except Exception:
                                            if log_callback: log_callback(
                                                f"⚠️ 所有格式复制方法都失败，跳过格式应用", False)
                                
                                # 清理剪贴板
                                try:
                                    import win32clipboard
                                    win32clipboard.OpenClipboard()
                                    win32clipboard.EmptyClipboard()
                                    win32clipboard.CloseClipboard()
                                except:
                                    # 如果清理剪贴板失败，忽略错误
                                    pass
                                
                            except Exception as e:
                                if log_callback: log_callback(
                                    f"警告：格式复制失败，数据已成功插入但无格式。原因: {e}", False)
                        else:
                            if log_callback: log_callback(
                                f"跳过格式复制：{'无模板行' if not template_row else '无数据行'}", False)
                            
                    except Exception as e:
                        if log_callback: log_callback(f"错误: 粘贴数据到工作表 '{target_sheet_name}' 失败。原因: {e}",
                                                      True)
                        if log_callback: log_callback(traceback.format_exc(), True)
                        continue
                else:
                    if log_callback: log_callback(f"没有数据需要粘贴到 '{target_sheet_name}'。", False)

        # Handle image insertion (Doc4)
        pic_sheet_name = '设备外观图'
        if doc4_path and os.path.exists(doc4_path):
            if log_callback: log_callback(f"\n正在处理图片文件 '{os.path.basename(doc4_path)}'", False)
            if pic_sheet_name in [s.name for s in wb.sheets]:
                pic_sht = wb.sheets[pic_sheet_name]
                if log_callback: log_callback(f"已找到工作表: '{pic_sheet_name}'", False)
            else:
                try:
                    pic_sht = wb.sheets.add(name=pic_sheet_name, after=wb.sheets[-1])
                    if log_callback: log_callback(f"创建工作表: '{pic_sheet_name}'", False)
                except Exception as e:
                    if log_callback: log_callback(f"错误: 无法创建图片工作表 '{pic_sheet_name}'。原因: {e}", True)
                    if log_callback: log_callback(traceback.format_exc(), True)
                    # Don't return, continue to save if other operations were successful
                    pic_sht = None  # Ensure pic_sht is None if creation failed

            if pic_sht:  # Only proceed if sheet exists or was created
                try:
                    second_row_top = pic_sht.range('2:2').top if pic_sht.used_range and pic_sht.used_range.api is not None else float('inf')
                    if log_callback: log_callback(
                        f"正在清除 '{pic_sheet_name}' 中第二行及以后所有图片...", False)
                    pictures_deleted_count = 0
                    for pic in list(pic_sht.pictures):
                        if pic.top >= second_row_top:
                            pic.delete()
                            pictures_deleted_count += 1
                    if log_callback: log_callback(f"已清除 {pictures_deleted_count} 张图片。", False)

                    width_pt = 23.66 * 28.3465
                    height_pt = 13.31 * 28.3465
                    top_left_cell = pic_sht.range('A2')

                    normalized_doc4_path = os.path.normpath(doc4_path)
                    if log_callback: log_callback(
                        f"准备插入图片。原始路径: '{doc4_path}', 规范化路径: '{normalized_doc4_path}'", False)
                    if log_callback: log_callback(
                        f"正在插入图片 '{os.path.basename(normalized_doc4_path)}' 到 '{pic_sheet_name}' 的 '{top_left_cell.address}'...",
                        False)
                    pic_sht.pictures.add(normalized_doc4_path,
                                         left=top_left_cell.left,
                                         top=top_left_cell.top,
                                         width=width_pt,
                                         height=height_pt)
                    if log_callback: log_callback("图片插入成功。", False)
                except Exception as e:
                    if log_callback: log_callback(f"错误: 插入图片到工作表 '{pic_sheet_name}' 失败。原因: {e}", True)
                    if log_callback: log_callback(traceback.format_exc(), True)
        else:
            if log_callback: log_callback(
                f"警告：图片文件 '{os.path.basename(doc4_path) if doc4_path else '未指定'}' 未选择或不存在，跳过图片插入。",
                False)

        wb.save()
        wb.close()
        if log_callback: log_callback("\n✅ 所有数据及图片已成功汇总到目标文件。", False)
        return True

    except Exception as e:
        if log_callback: log_callback(f"❌ 出现错误：{e}", True)
        if log_callback: log_callback("请确保：", True)
        if log_callback: log_callback("1. Microsoft Excel 已安装并可正常运行。", True)
        if log_callback: log_callback("2. 所有源文件和目标报告文件在操作过程中是关闭状态。", True)
        if log_callback: log_callback("3. 文件路径正确无误，且您有读写权限。", True)
        if log_callback: log_callback(
            "4. 目标工作表名称与配置一致（特别是 '遗留缺陷列表', '产品需求列表', '验收测试用例', '设备外观图'）。", True)
        if log_callback: log_callback(f"详细错误信息: {traceback.format_exc()}", True)
        return False
    finally:
        if app:
            try:
                app.quit()
                if log_callback: log_callback("xlwings 应用程序已关闭。", False)
            except Exception as cleanup_error:
                if log_callback: log_callback(f"清理xlwings应用程序时出错: {cleanup_error}", True)


def write_to_excel_with_xlwings(file_path, sheet_name, cell_map, data_dict, log_callback=None):
    import xlwings as xw
    app = None
    try:
        if log_callback: log_callback(f"使用xlwings打开文件: {file_path}")
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(file_path, update_links=False)
        if sheet_name not in [s.name for s in wb.sheets]:
            if log_callback: log_callback(f"找不到工作表: {sheet_name}", is_error=True)
            return False
        sht = wb.sheets[sheet_name]
        for key, cell in cell_map.items():
            value = data_dict.get(key, "")
            if value:
                sht.range(cell).value = value
                if log_callback: log_callback(f"写入 {key}: {value} 到 {cell}")
            else:
                if log_callback: log_callback(f"跳过 {key}，无内容", is_error=False)
        wb.save()
        if log_callback: log_callback("数据已成功写入并保存！")
        wb.close()
        return True
    except Exception as e:
        import traceback
        if log_callback: log_callback(f"xlwings写入异常: {e}", is_error=True)
        if log_callback: log_callback(traceback.format_exc(), is_error=True)
        return False
    finally:
        if app:
            try:
                app.quit()
            except Exception as cleanup_error:
                if log_callback: log_callback(f"清理xlwings应用程序时出错: {cleanup_error}", True)


def _consolidate_with_win32com_fallback(doc1_path: str, doc2_path: str, doc3_path: str, doc4_path: str,
                                       target_report_path: str, 
                                       append_mode_doc1: bool = False,
                                       append_mode_doc2: bool = False, 
                                       append_mode_doc3: bool = False,
                                       log_callback=None):
    """
    使用win32com作为备用方案进行数据汇总
    专门处理包含VBA代码的.xlsm文件，保持公式和数据交互功能
    """
    excel_app = None
    target_wb = None
    
    try:
        if log_callback: log_callback("使用win32com备用方案进行数据汇总（保持VBA功能）...", False)
        
        # 检查目标文件
        if not target_report_path or not os.path.exists(target_report_path):
            if log_callback: log_callback(f"错误：目标报告文件不存在: {target_report_path}", True)
            return False
        
        # 初始化Excel应用程序
        try:
            import win32com.client
            import pythoncom
            
            # 强制清理COM组件
            try:
                pythoncom.CoUninitialize()
            except:
                pass
            
            pythoncom.CoInitialize()
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            if log_callback: log_callback("win32com Excel应用程序初始化成功", False)
            
        except Exception as e:
            if log_callback: log_callback(f"win32com初始化失败: {e}", True)
            return False
        
        # 打开目标工作簿
        try:
            target_wb = excel_app.Workbooks.Open(target_report_path, UpdateLinks=False)
            if log_callback: log_callback(f"成功打开目标报告：{os.path.basename(target_report_path)}", False)
        except Exception as e:
            if log_callback: log_callback(f"无法打开目标报告文件: {e}", True)
            return False
        
        source_info = [
            {'path': doc1_path, 'sheet_name': '遗留缺陷列表', 'append_mode': append_mode_doc1},
            {'path': doc2_path, 'sheet_name': '产品需求列表', 'append_mode': append_mode_doc2},
            {'path': doc3_path, 'sheet_name': '验收测试用例', 'append_mode': append_mode_doc3}
        ]
        
        success_count = 0
        for info in source_info:
            src_path = info['path']
            target_sheet_name = info['sheet_name']
            append_mode = info['append_mode']
            
            if not src_path or not os.path.exists(src_path):
                if log_callback: log_callback(f"警告：源文件不存在，跳过: {src_path}", False)
                continue
            
            try:
                # 读取源数据
                df = pd.read_excel(src_path, header=None)
                if df.empty:
                    if log_callback: log_callback(f"警告：源文件为空: {src_path}", False)
                    continue
                
                data = df.iloc[1:].values.tolist()  # 跳过第一行
                
                # 处理遗留缺陷列表的过滤逻辑
                if target_sheet_name == '遗留缺陷列表' and data:
                    original_count = len(data)
                    bug_status_col_index = 9
                    filtered_data = []
                    
                    for row_index, row in enumerate(data):
                        if len(row) > bug_status_col_index:
                            bug_status = str(row[bug_status_col_index]).strip() if row[bug_status_col_index] else ""
                            if bug_status != "已关闭":
                                filtered_data.append(row)
                    
                    data = filtered_data
                    if log_callback: log_callback(f"过滤完成：原始{original_count}行，保留{len(data)}行", False)
                
                # 处理工作表
                target_sheet = None
                try:
                    target_sheet = target_wb.Worksheets(target_sheet_name)
                    if log_callback: log_callback(f"找到工作表: {target_sheet_name}", False)
                except:
                    # 创建工作表
                    try:
                        target_sheet = target_wb.Worksheets.Add(After=target_wb.Worksheets(target_wb.Worksheets.Count))
                        target_sheet.Name = target_sheet_name
                        if log_callback: log_callback(f"创建工作表: {target_sheet_name}", False)
                    except Exception as e:
                        if log_callback: log_callback(f"无法创建工作表: {target_sheet_name}, 错误: {e}", True)
                        continue
                
                if target_sheet:
                    # 确定起始行
                    start_row = 3
                    if append_mode:
                        # 追加模式：找到最后一行数据
                        last_row = target_sheet.UsedRange.Rows.Count
                        for row in range(start_row, last_row + 2):
                            # 检查该行是否有数据
                            has_data = False
                            for col in range(1, 6):
                                cell_value = target_sheet.Cells(row, col).Value
                                if cell_value is not None and str(cell_value).strip():
                                    has_data = True
                                    break
                            if not has_data:
                                start_row = row
                                break
                    
                    # 写入数据
                    for row_idx, row_data in enumerate(data):
                        for col_idx, cell_value in enumerate(row_data, 1):
                            target_sheet.Cells(start_row + row_idx, col_idx).Value = cell_value
                    
                    if log_callback: log_callback(f"成功写入{len(data)}行数据到{target_sheet_name}", False)
                    success_count += 1
                
            except Exception as e:
                if log_callback: log_callback(f"处理源文件失败: {src_path}, 错误: {e}", True)
                continue
        
        # 处理图片插入（如果有doc4_path）
        if doc4_path and os.path.exists(doc4_path):
            try:
                pic_sheet_name = '设备外观图'
                pic_sheet = None
                
                try:
                    pic_sheet = target_wb.Worksheets(pic_sheet_name)
                    if log_callback: log_callback(f"找到图片工作表: {pic_sheet_name}", False)
                except:
                    try:
                        pic_sheet = target_wb.Worksheets.Add(After=target_wb.Worksheets(target_wb.Worksheets.Count))
                        pic_sheet.Name = pic_sheet_name
                        if log_callback: log_callback(f"创建图片工作表: {pic_sheet_name}", False)
                    except Exception as e:
                        if log_callback: log_callback(f"无法创建图片工作表: {pic_sheet_name}, 错误: {e}", True)
                
                if pic_sheet:
                    # 清除第二行及以后的图片
                    try:
                        for shape in pic_sheet.Shapes:
                            if shape.Top >= pic_sheet.Rows(2).Top:
                                shape.Delete()
                        if log_callback: log_callback("已清除第二行及以后的图片", False)
                    except:
                        pass
                    
                    # 插入图片
                    try:
                        left = pic_sheet.Range("A2").Left
                        top = pic_sheet.Range("A2").Top
                        width = 23.66 * 28.3465  # 转换为磅
                        height = 13.31 * 28.3465
                        
                        pic_sheet.Shapes.AddPicture(
                            Filename=doc4_path,
                            LinkToFile=False,
                            SaveWithDocument=True,
                            Left=left,
                            Top=top,
                            Width=width,
                            Height=height
                        )
                        if log_callback: log_callback("图片插入成功", False)
                    except Exception as e:
                        if log_callback: log_callback(f"图片插入失败: {e}", True)
                        
            except Exception as e:
                if log_callback: log_callback(f"处理图片时出错: {e}", True)
        
        # 保存文件
        target_wb.Save()
        if log_callback: log_callback("文件保存成功", False)
        
        if success_count > 0:
            if log_callback: log_callback(f"✅ win32com备用方案完成，成功处理{success_count}个文件", False)
            return True
        else:
            if log_callback: log_callback("❌ win32com备用方案未成功处理任何文件", True)
            return False
            
    except Exception as e:
        if log_callback: log_callback(f"win32com备用方案执行失败: {e}", True)
        if log_callback: log_callback(traceback.format_exc(), True)
        return False
    finally:
        # 清理资源
        try:
            if target_wb:
                target_wb.Close(SaveChanges=False)
            if excel_app:
                excel_app.Quit()
            if log_callback: log_callback("win32com Excel应用程序已关闭", False)
        except Exception as cleanup_error:
            if log_callback: log_callback(f"清理win32com应用程序时出错: {cleanup_error}", True)
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass